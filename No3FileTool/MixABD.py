# -*- coding: UTF-8 -*-
from openpyxl import *
import os
import time
from datetime import datetime, timedelta

format_pattern = '%Y-%m-%d %H:%M:%S'
cur_time = datetime.now()
#当前时间减少2天
timeline = cur_time + timedelta(days=-2)
# 将 'cur_time' 类型时间通过格式化模式转换为 'str' 时间
timeline = timeline.strftime(format_pattern)

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print(113344422+len(file_lists))
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

def  match_add(ws3,ws1,ws2,ws4):
    allcol3 = ws3.max_column
    allrow3 = ws3.max_row
    allcol1 = ws1.max_column
    allrow1 = ws1.max_row
    allcol2 = ws2.max_column
    allrow2 = ws2.max_row
    allrow4 = ws4.max_row
    index_D_row = 2
    while index_D_row <= allrow3:  #循环D表的所有行
        
        found = False
        value_3 = ws3.cell(index_D_row,2).value #读取单号
        if  value_3 == None:
            index_D_row+=1
            continue
        else:
            value_3 = int(value_3)   #单号有效
            value_3_1 = ws3.cell(index_D_row,30).value    #读取地址
            value_3_2 = ws3.cell(index_D_row,40).value    #读取三段码
            value_3_3 = ws3.cell(index_D_row,8).value     #业务标签
            value_3_4 = ws3.cell(index_D_row,4).value     #获取时间
            value_3_4 = str(value_3_4)
            if value_3_3 != None:
                value_3_3 = str(value_3_3)
                value_3_3 = value_3_3[0:2]
            else:
                value_3_3 = 'not'
            if found == False:     #在A表中查找
                index_A_row = 2
                while index_A_row <= allrow1:
                    value_1 = ws1.cell(index_A_row,2).value
                    if value_1 == None:
                        index_A_row+=1
                        continue
                    else:
                        value_1 = int(value_1)
                        if value_1 == value_3:
                            found = True
                            if value_3_3 == '生鲜件':
                                ws1.cell(row = index_A_row,column = 12, value = value_3_3)      #标注‘生鲜件’
                            ws1.cell(row = index_A_row,column = 8, value = value_3_1)  #在A表中写入地址
                            if ws3.cell(index_D_row,11).value == '江苏盐城公司' or ws3.cell(index_D_row,11).value == '江苏省市场部五十七部':
                                difference = (datetime.strptime(value_3_4, format_pattern) - datetime.strptime(timeline, format_pattern))
                            if difference.days < 0:
                                ws1.cell(row = index_A_row,column = 9,value = value_3_2)
                                ws1.cell(row = index_A_row,column = 11,value = '退回件')
                            else:
                                if value_3_2 == None:             #如果三段码为空，直接写入
                                    ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                else:                 #如果三段码不为空，转换str，判断长度
                                    value_3_2 = str(value_3_2)
                                    if len(value_3_2)<9:            #三段码不全 直接写入
                                        ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                    else:                           #三段码完整 判断前两段是否属于盐城
                                        if value_3_2[0:3] != '466' and value_3_2[0:3]!='467': #第一段不属于盐城，直接写入
                                            ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                        else:
                                            if value_3_2[4:7]!='001' and value_3_2[4:7]!='AA1': #第二段不属于盐城，直接写入
                                                ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                            else:                         #完全属于盐城，直接截取最后一段写入
                                                ws1.cell(row = index_A_row,column = 9, value = value_3_2[-3:])
                                                index_3 = 3
                                                while index_3 <= allrow4:    #查找对应的展示名称 并写入
                                                    value4 = ws4.cell(index_3,1).value
                                                    value4 = str(value4)
                                                    if value4 == value_3_2[-3:]:
                                                        ws1.cell(row = index_A_row , column = 10 , value = ws4.cell(index_3,2).value)
                                                        break
                                                    else:
                                                        index_3+=1
                            else:
                                if value_3_2 == None:             #如果三段码为空，直接写入
                                    ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                else:                               #如果三段码不为空，转换str，判断长度
                                    value_3_2 = str(value_3_2)
                                    if len(value_3_2)<9:            #三段码不全 直接写入
                                        ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                    else:                           #三段码完整 判断前两段是否属于盐城
                                        if value_3_2[0:3] != '466' and value_3_2[0:3]!='467': #第一段不属于盐城，直接写入
                                            ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                        else:
                                            if value_3_2[4:7]!='001' and value_3_2[4:7]!='AA1': #第二段不属于盐城，直接写入
                                                ws1.cell(row = index_A_row,column = 9, value = value_3_2)
                                            else:                         #完全属于盐城，直接截取最后一段写入
                                                ws1.cell(row = index_A_row,column = 9, value = value_3_2[-3:])
                                                index_3 = 3
                                                while index_3 <= allrow4:    #查找对应的展示名称 并写入
                                                    value4 = ws4.cell(index_3,1).value
                                                    value4 = str(value4)
                                                    if value4 == value_3_2[-3:]:
                                                        ws1.cell(row = index_A_row , column = 10 , value = ws4.cell(index_3,2).value)
                                                        break
                                                    else:
                                                        index_3+=1
                            break
                        else:
                            index_A_row+=1
            if found == False:    #在B表中查找
                index_B_row = 2
                while index_B_row <= allrow2:
                    value_2 = ws2.cell(index_B_row,1).value
                    if value_2 == None:
                        index_B_row+=1
                        continue
                    else:
                        value_2 = int(value_2)
                        if value_2 == value_3:
                            found = True
                            if value_3_3 == '生鲜件':
                                ws2.cell(row = index_B_row,column = 8,value = value_3_3)    #标注‘生鲜件’ws2.cell(row = index_B_row,column = 3,value = value_3_1)    #在B表中写入地址
							if ws3.cell(index_D_row,11).value == '江苏盐城公司' or ws3.cell(index_D_row,11).value == '江苏省市场部五十七部':
								difference = (datetime.strptime(value_3_4, format_pattern) - datetime.strptime(timeline, format_pattern))
								if difference.days < 0:
									ws2.cell(row = index_B_row,column = 4,value = value_3_2)
									ws2.cell(row = index_B_row,column = 7,value = '退回件')
								else:
									if value_3_2 == None:
										ws2.cell(row = index_B_row,column = 4,value = value_3_2)
									else:
										value_3_2 = str(value_3_2)
										if len(value_3_2)<9:
											ws2.cell(row = index_B_row,column = 4,value = value_3_2)
										else:
											if value_3_2[0:3] != '466' and value_3_2[0:3]!='467':
												ws2.cell(row = index_B_row,column = 4,value = value_3_2)
											else:
												if value_3_2[4:7]!='001' and value_3_2[4:7]!='AA1':
													ws2.cell(row = index_B_row,column = 4,value = value_3_2)
												else:
													ws2.cell(row = index_B_row,column = 4,value = value_3_2[-3:])
													index_3=3
													while index_3 <= allrow4:
														value4 = ws4.cell(index_3,1).value
														value4 = str(value4)
														if value4 == value_3_2[-3:]:
															ws2.cell(row = index_B_row,column = 5,value = ws4.cell(index_3,2).value)
															break
														else:
															index_3+=1
                            else:
                                if value_3_2 == None:
                                    ws2.cell(row = index_B_row,column = 4,value = value_3_2)
                                else:
                                    value_3_2 = str(value_3_2)
                                    if len(value_3_2)<9:
                                        ws2.cell(row = index_B_row,column = 4,value = value_3_2)
                                    else:
                                        if value_3_2[0:3] != '466' and value_3_2[0:3]!='467':
                                            ws2.cell(row = index_B_row,column = 4,value = value_3_2)
                                        else:
                                            if value_3_2[4:7]!='001' and value_3_2[4:7]!='AA1':
                                                ws2.cell(row = index_B_row,column = 4,value = value_3_2)
                                            else:
                                                ws2.cell(row = index_B_row,column = 4,value = value_3_2[-3:])
                                                index_3=3
                                                while index_3 <= allrow4:
                                                    value4 = ws4.cell(index_3,1).value
                                                    value4 = str(value4)
                                                    if value4 == value_3_2[-3:]:
                                                        ws2.cell(row = index_B_row,column = 5,value = ws4.cell(index_3,2).value)
                                                        break
                                                    else:
                                                        index_3+=1
                            break
                        else:
                            index_B_row+=1
            index_D_row+=1


readA_path = "/var/www/html/QualityCtrl/No1FileTool/MixAB_A/"
readB_path = "/var/www/html/QualityCtrl/No1FileTool/MixAB_B/"
readD_path = "/var/www/html/QualityCtrl/No3FileTool/uploadD/"
read3_path = "/var/www/html/QualityCtrl/No3FileTool/3codelist/"
save_path = "/var/www/html/QualityCtrl/No3FileTool/MixABD/" #用来保存文件 的 路径

file_name_A = find_new_file(readA_path)
file_name_B = find_new_file(readB_path)
file_name_D = find_new_file(readD_path)
file_name_3 = find_new_file(read3_path)


#业务逻辑
wb1 = load_workbook(readA_path+file_name_A)
ws1 = wb1[wb1.sheetnames[0]]
wb2 = load_workbook(readB_path+file_name_B)
ws2 = wb2[wb2.sheetnames[0]]
wb3 = load_workbook(readD_path+file_name_D)
ws3 = wb3[wb3.sheetnames[0]]
wb4 = load_workbook(read3_path+file_name_3)
ws4 = wb4[wb4.sheetnames[0]]

match_add(ws3,ws1,ws2,ws4)

wb1.save(save_path+file_name_A)
wb2.save(save_path+file_name_B)
wb3.save(readD_path+file_name_D)
wb4.save(read3_path+file_name_3)
