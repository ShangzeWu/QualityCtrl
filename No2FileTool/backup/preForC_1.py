# -*- coding: UTF-8 -*-
from openpyxl import *
import os

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/QualityCtrl/No2FileTool"
#print(path)
dir_C = path+'/uploadC/' #用来读取C文件 的 路径
dir_namelist = path+'/namelist/namelist.xlsx' #用来读取人员名单表 的 路径
dir_template = path+'/template/template.xlsx' #用来读取输出模版表格 的 路径
dir_save_C= "/var/www/html/QualityCtrl/No2FileTool/"  #输出 C文件 的保存路径

file_name_C = find_new_file(dir_C)


#业务逻辑
wb1 = load_workbook(dir_C+file_name_C) #C表
ws1 = wb1[wb1.sheetnames[0]]           #C表第一页
#wb2 = load_workbook(dir_namelist)      #namelist表
#ws2 = wb2[wb2.sheetnames[0]]           #namelist表第一页
#wb3 = load_workbook(dir_template)      #模板表

#去重
index_rm_row = 2
All_rm_row = ws1.max_row
while index_rm_row<=All_rm_row:
    name_runner=ws1.cell(index_rm_row,6).value  #读取派件员的名字
    if name_runner == None: #读到了空行
        index_rm_row+=1
        continue
    else:  # ！空行
        name_runner=str(name_runner)
        index_rm_row_inner = index_rm_row+1
        while index_rm_row_inner <= All_rm_row:
            value_inter1 = int(ws1.cell(index_rm_row,10).value) #存储已签收
            value_inter2 = int(ws1.cell(index_rm_row,12).value) #存储已派未签
            name_runner_inner = ws1.cell(index_rm_row_inner,6).value
            if name_runner_inner ==None:
                index_rm_row_inner+=1
                continue
            else:
                name_runner_inner = str(name_runner_inner)
                if name_runner_inner == name_runner:
                    value_inter1 = value_inter1+int(ws1.cell(index_rm_row_inner,10).value)
                    value_inter2 = value_inter2+int(ws1.cell(index_rm_row_inner,12).value)
                    ws1.cell(row=index_rm_row,column=10,value=value_inter1)
                    ws1.cell(row=index_rm_row,column=12,value=value_inter2)
                    ws1.delete_rows(index_rm_row_inner,1)
                    index_rm_row_inner= index_rm_row_inner-1
                index_rm_row_inner+=1
        index_rm_row+=1

wb1.save(dir_save_C+"resultC/"+file_name_C)

