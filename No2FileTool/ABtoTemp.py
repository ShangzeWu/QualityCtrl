# -*- coding: UTF-8 -*-
from openpyxl import *
import os

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

def classify(string):		#查找派件员所属大区
	if string == '益林区滞留':
		return 11
	if string == '运营部朱华荣':
		return 3
	if string =='机动人员施凯迪':
		return 4
	str1 = string[0:2] #获得字符的前两位
	if str1 == '宝龙':
		return 1
	elif str1 == '新都':
		return 2
	elif str1 == '城南':
		return 3
	elif str1 == '开发':
		return 4
	elif str1 == '五星':
		return 5
	elif str1 == '亭湖':
		return 6
	elif str1 == '吾悦':
		return 7
	elif str1 == 'LG':
		return 8
	elif str1 == '城北':
		return 9
	elif str1 == '盐都':
		return 10
	elif str1 == '乡镇':
		return 11
	elif str1 == '万达':
		return 12
	elif str1 == '招商':
		return 13
	else:
		return 14

def find_insert(ws,name,valueA_5):  #在工作表中查找是否存在这个人员
	ws_rows = ws.max_row
	index = 6
	while index<= ws_rows:
		ws_name = ws.cell(index,1).value
		if ws_name == None:
			index+=1
			continue
		else:
			ws_name = str(ws_name)
			if name == ws_name:    #发现存在该人员，总计数加1
				cur_count = ws.cell(index,6).value
				if cur_count == None:
					ws.cell(row = index, column = 6, value = 1)
				else:
					cur_count = int(cur_count)
					ws.cell(row = index, column = 6, value = cur_count+1)

				if valueA_5 == None: 			#漏扫件加1
					cur_count1 = ws.cell(index,8).value
					if cur_count1 == None:
						ws.cell(row = index, column = 8, value = 1)
					else:
						cur_count1 = int(cur_count1)
						ws.cell(row = index, column = 8, value = cur_count1+1)
					return
				else:
					cur_count2 = ws.cell(index,7).value
					if cur_count2 == None:
						ws.cell(row = index, column = 7, value = 1)
					else:
						cur_count2 = int(cur_count2)
						ws.cell(row = index, column = 7, value = cur_count2+1)
					return
			else:   #寻找下一行
				index+=1

def find_insert_b(ws,name):
	ws_rows = ws.max_row
	index = 6
	while index<= ws_rows:
		ws_name = ws.cell(index,1).value
		if ws_name == None:
			index+=1
			continue
		else:
			ws_name = str(ws_name)
			if name == ws_name:    #发现存在该人员，总计数加1
				cur_count = ws.cell(index,9).value
				if cur_count == None:
					ws.cell(row = index, column = 9, value = 1)
				else:
					cur_count = int(cur_count)
					ws.cell(row = index, column = 9, value = cur_count+1)
				return
			else:   #寻找下一行
				index+=1

def add(ws_a,ws):  #单表合计
	ws_row = ws.max_row
	counter1 = 0 #计数器
	counter2 = 0
	counter3 = 0
	counter4 = 0
	index = 6 #起始行数
	while index <= ws_row-2:
		value_temp1 = ws.cell(index,6).value
		value_temp2 = ws.cell(index,7).value
		value_temp3 = ws.cell(index,8).value
		value_temp4 = ws.cell(index,9).value
		if value_temp1 == None and value_temp4 == None:   #总量为0，直接下一行
			index+=1
			continue
		elif value_temp1!= None and value_temp4 == None:  #到件无下文！=0，发往无下文==0
			value_temp1 = int(value_temp1)
			counter1 = counter1+value_temp1
			if value_temp2 != None:
				value_temp2 = int(value_temp2)
				counter2 = counter2+value_temp2
			if value_temp3 != None:
				value_temp3 = int(value_temp3)
				counter3 = counter3+value_temp3
			index+=1
		elif value_temp1 == None and value_temp4 != None:  #到件无下文==0，发往无下文！=0
			value_temp4 = int(value_temp4)
			counter4 = counter4+value_temp4
			index+=1
		elif value_temp1 != None  and value_temp4 !=None:  #到件无下文！=0，发往无下文！=0
			value_temp4 = int(value_temp4)
			counter4 = counter4+value_temp4
			value_temp1 = int(value_temp1)
			counter1 = counter1+value_temp1
			if value_temp2 != None:
				value_temp2 = int(value_temp2)
				counter2 = counter2+value_temp2
			if value_temp3 != None:
				value_temp3 = int(value_temp3)
				counter3 = counter3+value_temp3
			index+=1

	ws.cell(row=ws_row-1,column=6,value=counter1)
	ws.cell(row=ws_row-1,column=7,value=counter2)
	ws.cell(row=ws_row-1,column=8,value=counter3)
	ws.cell(row=ws_row-1,column=9,value=counter4)
	#添加进总表
	string = ws.title
	string = str(string)
	if string == '城南区':
		ws_a.cell(row=17,column=7,value=counter1)
		ws_a.cell(row=17,column=8,value=counter2)
		ws_a.cell(row=17,column=9,value=counter3)
		ws_a.cell(row=17,column=10,value=counter4)
	for i in range(5,19):
		name_of_zone = ws_a.cell(i,2).value
		if name_of_zone == string:
			ws_a.cell(row=i,column=7,value=counter1)
			ws_a.cell(row=i,column=8,value=counter2)
			ws_a.cell(row=i,column=9,value=counter3)
			ws_a.cell(row=i,column=10,value=counter4)



path =  "/var/www/html/QualityCtrl/No2FileTool"
dir_template = path+'/resultTemp/temp.xlsx' #用来读取已经写入过C的汇总表 的 路径
dir_MixABD= "/var/www/html/QualityCtrl/No2FileTool/uploadMixABD/"  #输出 C文件 的保存路径

file_name_ABD = find_new_file(dir_MixABD)


#业务逻辑
wb1 = load_workbook(dir_MixABD+file_name_ABD) 		#处理后的ABD混合输入表
ws1_a = wb1['a 到件无下文']				#ABD表中的A表
ws1_b = wb1['b发往无下文']          		#ABD表中的B表
wb3 = load_workbook(dir_template)      #模板表
ws3_all = wb3['汇总22.00']      #已经写入过C的汇总表
ws3_BL  = wb3['宝龙区']
ws3_XD  = wb3['新都区']
ws3_CN  = wb3['城南区']
ws3_KF  = wb3['开发区']
ws3_WX  = wb3['五星区']
ws3_TH  = wb3['亭湖区']
ws3_WY  = wb3['吾悦区']
ws3_LG  = wb3['龙冈区']
ws3_CB  = wb3['城北区']
ws3_YD  = wb3['盐都区']
ws3_YL  = wb3['益林区']
ws3_WD  = wb3['万达区']
ws3_ZS  = wb3['招商区']
ws3_other  = wb3['其他']

#print(ws3_BL.title)
#print(ws3_XD.max_row)
#print(ws3_CN.max_row)

#wb.load_workbook(dir_save_C+file_name_C)
#ws.wb[wb.sheetnames[0]]
AllrowA = ws1_a.max_row
AllrowB = ws1_b.max_row

index = 3
while index<=AllrowA:
	name = ws1_a.cell(index,3).value  #获得ABD表中A表的操作人员名
	if name == None:
		index += 1
		continue
	else:
		name = str(name)
		valueA_5 = ws1_a.cell(index,5).value 	 	#检查是否发往下一站/二次到件
		index_sheet = classify(name) #获取对应的工作页索引
		if index_sheet == 1:
			find_insert(ws3_BL,name,valueA_5)
		elif index_sheet == 2:
			find_insert(ws3_XD,name,valueA_5)
		elif index_sheet == 3:
			find_insert(ws3_CN,name,valueA_5)
		elif index_sheet == 4:
			find_insert(ws3_KF,name,valueA_5)
		elif index_sheet == 5:
			find_insert(ws3_WX,name,valueA_5)
		elif index_sheet == 6:
			find_insert(ws3_TH,name,valueA_5)
		elif index_sheet == 7:
			find_insert(ws3_WY,name,valueA_5)
		elif index_sheet == 8:
			find_insert(ws3_LG,name,valueA_5)
		elif index_sheet == 9:
			find_insert(ws3_CB,name,valueA_5)
		elif index_sheet == 10:
			find_insert(ws3_YD,name,valueA_5)
		elif index_sheet == 11:
			find_insert(ws3_YL,name,valueA_5)
		elif index_sheet == 12:
			find_insert(ws3_WD,name,valueA_5)
		elif index_sheet == 13:
			find_insert(ws3_ZS,name,valueA_5)
		else:
			find_insert(ws3_other,name,valueA_5)
#		print(index_sheet)
		index += 1

index_b = 3
while index_b<=AllrowB:
	name_b = ws1_b.cell(index_b,2).value
	if name_b == None:
		index_b+=1
		continue
	else:
		name_b = str(name_b)
		index_sheet_b = classify(name_b)
		if index_sheet_b == 1:
			find_insert_b(ws3_BL,name_b)
		elif index_sheet_b == 2:
			find_insert_b(ws3_XD,name_b)
		elif index_sheet_b == 3:
			find_insert_b(ws3_CN,name_b)
		elif index_sheet_b == 4:
			find_insert_b(ws3_KF,name_b)
		elif index_sheet_b == 5:
			find_insert_b(ws3_WX,name_b)
		elif index_sheet_b == 6:
			find_insert_b(ws3_TH,name_b)
		elif index_sheet_b == 7:
			find_insert_b(ws3_WY,name_b)
		elif index_sheet_b == 8:
			find_insert_b(ws3_LG,name_b)
		elif index_sheet_b == 9:
			find_insert_b(ws3_CB,name_b)
		elif index_sheet_b == 10:
			find_insert_b(ws3_YD,name_b)
		elif index_sheet_b == 11:
			find_insert_b(ws3_YL,name_b)
		elif index_sheet_b == 12:
			find_insert_b(ws3_WD,name_b)
		elif index_sheet_b == 13:
			find_insert_b(ws3_ZS,name_b)
		else:
			find_insert_b(ws3_other,name_b)
		index += 1

add(ws3_all,ws3_BL)
add(ws3_all,ws3_XD)
add(ws3_all,ws3_CN)
add(ws3_all,ws3_KF)
add(ws3_all,ws3_WX)
add(ws3_all,ws3_TH)
add(ws3_all,ws3_WY)
add(ws3_all,ws3_LG)
add(ws3_all,ws3_CB)
add(ws3_all,ws3_YD)
add(ws3_all,ws3_YL)
add(ws3_all,ws3_WD)
add(ws3_all,ws3_ZS)
add(ws3_all,ws3_other)

counterA=0
counterB=0
counterC=0
counterD=0
for i in range(5,19):
	valueA = ws3_all.cell(i,7).value
	valueB = ws3_all.cell(i,8).value
	valueC = ws3_all.cell(i,9).value
	valueD = ws3_all.cell(i,10).value
	counterA = counterA+valueA
	if valueB != None:
		counterB = counterB+valueB
	if valueC != None:
		counterC = counterC+valueC
	if valueD != None:
		counterD = counterD+valueD

ws3_all.cell(row=19,column=7,value=counterA)
ws3_all.cell(row=19,column=8,value=counterB)
ws3_all.cell(row=19,column=9,value=counterC)
ws3_all.cell(row=19,column=10,value=counterD)

wb1.save(dir_MixABD+file_name_ABD)
wb3.save(dir_template)
