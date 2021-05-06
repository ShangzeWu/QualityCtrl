# -*- coding: UTF-8 -*-
from openpyxl import *
import os

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/QualityCtrl/No2FileTool"
#print(path)
dir_C = path+'/uploadC/' #用来读取C文件 的 路径
dir_namelist = path+'/namelist/namelist.xlsx' #用来读取人员名单表 的 路径
#dir_template = path+'/template/template.xlsx' #用来读取输出模版表格 的 路径
dir_save_C= "/var/www/html/QualityCtrl/No2FileTool/"  #输出 C文件 的保存路径

file_name_C = find_new_file(dir_C)

#去除空行
wb3 = load_workbook(dir_save_C+"resultC/"+file_name_C)
ws3 = wb3[wb3.sheetnames[0]]
Allrow1_new = ws3.max_row
#print(Allrow1_new)
#print(ws3.cell(26,6).value)
counter = 1
index_rm_void = 2
while index_rm_void<=Allrow1_new and counter <=Allrow1_new:
    counter+=1
    void_runner = ws3.cell(index_rm_void,6).value
    if void_runner == None:
        ws3.delete_rows(index_rm_void,1)
        print(index_rm_void)
        index_rm_void-=1
    index_rm_void+=1
wb3.save(dir_save_C+"resultC/"+file_name_C)
