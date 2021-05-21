# -*- coding: UTF-8 -*-
from openpyxl import *
import os
import time
from datetime import datetime, timedelta
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment

format_pattern = '%Y-%m-%d %H:%M:%S'
cur_time = datetime.now()
cur_time1 = datetime.now()
# 将 'cur_time' 类型时间通过格式化模式转换为 'str' 时间
cur_time = cur_time.strftime(format_pattern)
str_name1 = '何红'
str_name2 = '李晓萍'
teleNum1  = 18921880822
teleNum2  = 17751570135

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

def classify(string):
	if string == '益林区滞留':
		return 11
	if string == '运营部朱华荣':
		return 3
	if string =='机动人员施凯迪':
		return 4
	str1 = string[0:2] #获得字符的前两位
	if str1 == '宝龙':
		return 1
	elif str1 == '新都':
		return 2
	elif str1 == '城南':
		return 3
	elif str1 == '开发':
		return 4
	elif str1 == '五星':
		return 5
	elif str1 == '亭湖':
		return 6
	elif str1 == '吾悦':
		return 7
	elif str1 == 'LG':
		return 8
	elif str1 == '城北':
		return 9
	elif str1 == '盐都':
		return 10
	elif str1 == '乡镇':
		return 11
	elif str1 == '万达':
		return 12
	elif str1 == '招商':
		return 13
	else:
		return 14

def find_insert(ws,name,value10,value12):  #在工作表中查找是否存在这个人员
	ws_rows = ws.max_row
	index = 6
	while index<= ws_rows:
		ws_name = ws.cell(index,1).value
		if ws_name == None:
			index+=1
			continue
		else:
			ws_name = str(ws_name)
			if name == ws_name:    #发现存在该人员，插入数据
				ws.cell(row=index,column=3,value=value10)
				ws.cell(row=index,column=5,value=value12)
				return
			else:   #寻找下一行
				index+=1
	if index>ws_rows:  #遍历工作页，没找到对应人员，插入一行并写入数据
		ws.insert_rows(6,1)
		font = ws["B"+str(index-3)].font
		ws.cell(row=6,column=1,value=name)
		ws.cell(row=6,column=3,value=value10)
		ws.cell(row=6,column=5,value=value12)
		ws['D6'].number_format = '0.00%'
#		print(font)
		ws['A6'].font = Font(name=u'微软雅黑', size=10)
		ws['B6'].font = Font(name=u'微软雅黑', size=10)
		ws['C6'].font = Font(name=u'微软雅黑', size=10)
		ws['D6'].font = Font(name=u'微软雅黑', size=10)
		ws['E6'].font = Font(name=u'微软雅黑', size=10)
		ws['F6'].font = Font(name=u'微软雅黑', size=10)
		ws['G6'].font = Font(name=u'微软雅黑', size=10)
		ws['H6'].font = Font(name=u'微软雅黑', size=10)
		ws['I6'].font = Font(name=u'微软雅黑', size=10)
		ws['J6'].font = Font(name=u'微软雅黑', size=10)
		ws['K6'].font = Font(name=u'微软雅黑', size=10)
		ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['D6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['E6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['F6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['G6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['H6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['I6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['J6'].alignment = Alignment(horizontal='center', vertical='center')
		ws['K6'].alignment = Alignment(horizontal='center', vertical='center')
		return

def add(ws_a,ws):  #单表合计
	ws.cell(row = 2,column = 5 , value = cur_time)
	ws_row = ws.max_row
	counter1 = 0 #计数器
	counter2 = 0
	index = 6 #起始行数
	while index <= ws_row-2:
		value_temp1 = ws.cell(index,3).value
		value_temp2 = ws.cell(index,5).value
		if value_temp1 == None and value_temp2 == None:
			index+=1
			continue
		else:
			value_temp1 = int(value_temp1)
			value_temp2 = int(value_temp2)
			counter1 = counter1+value_temp1
			counter2 = counter2+value_temp2
			index+=1
	ws.cell(row=ws_row-1,column=3,value=counter1)
	ws.cell(row=ws_row-1,column=5,value=counter2)
	string = ws.title
	string = str(string)
	for i in range(5,19):
		name_of_zone = ws_a.cell(i,2).value
		if name_of_zone == string:
			ws_a.cell(row=i,column=4,value=counter1)
			ws_a.cell(row=i,column=6,value=counter2)
		if string == '城南区':
			ws_a.cell(row=17,column=4,value=counter1)
			ws_a.cell(row=17,column=6,value=counter2)

def inner_add(ws): #计算分表的合计和签收率
	ws_row = ws.max_row
#	ws_col = ws.max_column
	index = 6
	ws.auto_filter.ref = "A5:K"+str(ws_row-2)
#	ws.auto_filter.add_sort_condition("D6:D"+str(ws_row-2))
	while index<=ws_row-2:
		value_receive = ws.cell(index,3).value
		value_notsign = ws.cell(index,5).value
		if value_receive == None and value_notsign == None:
			index+=1
			continue
		else:
			value_receive = int(value_receive)
			value_notsign = int(value_notsign)
			value_all = value_receive + value_notsign
			ws.cell(row=index,column=2,value = value_all)
			ws.cell(row=index,column=10,value = value_notsign)
			ws.cell(row=index,column=4,value = value_receive/value_all)
			index+=1
			
		

path =  "/var/www/html/QualityCtrl/No2FileTool"
dir_template = path+'/template/' #用来读取输出模版表格 的 路径
dir_save_C= "/var/www/html/QualityCtrl/No2FileTool/resultC/"  #输出 C文件 的保存路径

file_name_C = find_new_file(dir_save_C)
file_name_temp = find_new_file(dir_template)


#业务逻辑
wb1 = load_workbook(dir_save_C+file_name_C) #处理后的C表
ws1 = wb1[wb1.sheetnames[0]]           #C表第一页
wb3 = load_workbook(dir_template+file_name_temp)      #模板表
ws3_all = wb3['汇总22.00']      #汇总表
ws3_BL  = wb3['宝龙区']
ws3_XD  = wb3['新都区']
ws3_CN  = wb3['城南区']
ws3_KF  = wb3['开发区']
ws3_WX  = wb3['五星区']
ws3_TH  = wb3['亭湖区']
ws3_WY  = wb3['吾悦区']
ws3_LG  = wb3['龙冈区']
ws3_CB  = wb3['城北区']
ws3_YD  = wb3['盐都区']
ws3_YL  = wb3['益林区']
ws3_WD  = wb3['万达区']
ws3_ZS  = wb3['招商区']
ws3_other  = wb3['其他']

#print(ws3_BL.title)
#print(ws3_XD.max_row)
#print(ws3_CN.max_row)

#wb.load_workbook(dir_save_C+file_name_C)
#ws.wb[wb.sheetnames[0]]
Allrow = ws1.max_row
#ws3_all.cell(row = 2,column = 5 , value = cur_time)

index = 2
while index<=Allrow:
	name = ws1.cell(index,6).value
	if name == None:
		index += 1
		continue
	else:
		name = str(name)
		value10 = ws1.cell(index,10).value
		value12 = ws1.cell(index,12).value
		value10 = int(value10)
		value12 = int(value12)
		index_sheet = classify(name) #获取对应的工作页索引
		if index_sheet == 1:
			find_insert(ws3_BL,name,value10,value12)
		elif index_sheet == 2:
			find_insert(ws3_XD,name,value10,value12)
		elif index_sheet == 3:
			find_insert(ws3_CN,name,value10,value12)
		elif index_sheet == 4:
			find_insert(ws3_KF,name,value10,value12)
		elif index_sheet == 5:
			find_insert(ws3_WX,name,value10,value12)
		elif index_sheet == 6:
			find_insert(ws3_TH,name,value10,value12)
		elif index_sheet == 7:
			find_insert(ws3_WY,name,value10,value12)
		elif index_sheet == 8:
			find_insert(ws3_LG,name,value10,value12)
		elif index_sheet == 9:
			find_insert(ws3_CB,name,value10,value12)
		elif index_sheet == 10:
			find_insert(ws3_YD,name,value10,value12)
		elif index_sheet == 11:
			find_insert(ws3_YL,name,value10,value12)
		elif index_sheet == 12:
			find_insert(ws3_WD,name,value10,value12)
		elif index_sheet == 13:
			find_insert(ws3_ZS,name,value10,value12)
		else:
			find_insert(ws3_other,name,value10,value12)
#		print(index_sheet)
		index += 1

add(ws3_all,ws3_BL)
add(ws3_all,ws3_XD)
add(ws3_all,ws3_CN)
add(ws3_all,ws3_KF)
add(ws3_all,ws3_WX)
add(ws3_all,ws3_TH)
add(ws3_all,ws3_WY)
add(ws3_all,ws3_LG)
add(ws3_all,ws3_CB)
add(ws3_all,ws3_YD)
add(ws3_all,ws3_YL)
add(ws3_all,ws3_WD)
add(ws3_all,ws3_ZS)
add(ws3_all,ws3_other)

inner_add(ws3_BL)
inner_add(ws3_XD)
inner_add(ws3_CN)
inner_add(ws3_KF)
inner_add(ws3_WX)
inner_add(ws3_TH)
inner_add(ws3_WY)
inner_add(ws3_LG)
inner_add(ws3_CB)
inner_add(ws3_YD)
inner_add(ws3_YL)
inner_add(ws3_WD)
inner_add(ws3_ZS)
inner_add(ws3_other)

counterA=0
counterB=0
for i in range(5,19):
	valueA = ws3_all.cell(i,4).value
	valueB = ws3_all.cell(i,6).value
	counterA = counterA+valueA
	counterB = counterB+valueB

ws3_all.cell(row=19,column=4,value=counterA)
ws3_all.cell(row=19,column=6,value=counterB)

wb1.save(dir_save_C+file_name_C)
wb3.save(path+'/resultTemp/Temp'+cur_time+'.xlsx')
