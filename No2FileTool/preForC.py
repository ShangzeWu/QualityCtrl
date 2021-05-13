# -*- coding: UTF-8 -*-
from openpyxl import *
import os
import time
#from datetime import datetime, timedelta

format_pattern = '%Y-%m-%d %H:%M:%S'
cur_time = datetime.now()
# 将 'cur_time' 类型时间通过格式化模式转换为 'str' 时间
cur_time = cur_time.strftime(format_pattern)

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/QualityCtrl/No2FileTool"
#print(path)
dir_C = path+'/uploadC/' #用来读取C文件 的 路径
dir_namelist = path+'/namelist/' #用来读取人员名单表 的 路径
#dir_template = path+'/template/template.xlsx' #用来读取输出模版表格 的 路径
dir_save_C= "/var/www/html/QualityCtrl/No2FileTool/"  #输出 C文件 的保存路径

file_name_C = find_new_file(dir_C)
file_name_list = find_new_file(dir_namelist)


#业务逻辑
wb1 = load_workbook(dir_C+file_name_C) #C表
ws1 = wb1[wb1.sheetnames[0]]           #C表第一页
wb2 = load_workbook(dir_namelist+file_name_list)      #namelist表
ws2 = wb2[wb2.sheetnames[0]]           #namelist表第一页
#wb3 = load_workbook(dir_template)      #模板表

#去重
index_rm_row = 2
All_rm_row = ws1.max_row
while index_rm_row<=All_rm_row:
    name_runner=ws1.cell(index_rm_row,6).value  #读取派件员的名字
    if name_runner == None: #读到了空行
        index_rm_row+=1
        continue
    else:  # ！空行
        name_runner=str(name_runner)
        index_rm_row_inner = index_rm_row+1
        while index_rm_row_inner <= All_rm_row:
            value_inter1 = int(ws1.cell(index_rm_row,10).value) #存储已签收
            value_inter2 = int(ws1.cell(index_rm_row,12).value) #存储已派未签
            name_runner_inner = ws1.cell(index_rm_row_inner,6).value
            if name_runner_inner ==None:
                index_rm_row_inner+=1
                continue
            else:
                name_runner_inner = str(name_runner_inner)
                if name_runner_inner == name_runner:
                    value_inter1 = value_inter1+int(ws1.cell(index_rm_row_inner,10).value)
                    value_inter2 = value_inter2+int(ws1.cell(index_rm_row_inner,12).value)
                    ws1.cell(row=index_rm_row,column=10,value=value_inter1)
                    ws1.cell(row=index_rm_row,column=12,value=value_inter2)
                    ws1.delete_rows(index_rm_row_inner,1)
                    index_rm_row_inner= index_rm_row_inner-1
                index_rm_row_inner+=1
        index_rm_row+=1

#合并三个表
Allrow1 = ws1.max_row
Allcol1 = ws1.max_column
Allrow2 = ws2.max_row
Allcol2 = ws2.max_column
#print(Allrow1)
index_C_col=1
while index_C_col<=Allcol1:
        if ws1.cell(1,index_C_col).value=="派件员":
            index_C_row = 2
            while index_C_row<=Allrow1:
                name_C = ws1.cell(index_C_row,index_C_col).value
                flag = True
                if name_C == None:  #排除C表的空行
                    index_C_row += 1
                    continue
                else:
                    name_C = str(name_C)
                    index_N_col=1
                    index_N_row=2
                    while index_N_row<=Allrow2:
                        name_N = ws2.cell(index_N_row,index_N_col).value
                        if name_N == None:
                            index_N_row+=1
                            continue
                        else:
                            name_N = str(name_N)
                            if name_N == name_C:
                                break
                            else:
                                index_N_row+=1
                        if index_N_row == Allrow2+1: #最后一次内层循环
                            flag = False   #没有找到C表要找的人
                    if flag==False:
                        ws1.delete_rows(index_C_row,1)
                        index_C_row = index_C_row-1
                    index_C_row+=1
        index_C_col=index_C_col+1
wb1.save(dir_save_C+"resultC/ChangedC"+cur_time+'.xlsx')
wb2.save(dir_namelist+file_name_list)
'''#去除空行
wb3 = load_workbook(dir_save_C+"resultC/"+file_name_C)
ws3 = wb3[wb3.sheetnames[0]]
Allrow1_new = ws3.max_row
#print(Allrow1_new)
print(ws3.cell(26,6).value)
counter = 1
index_rm_void = 2
while index_rm_void<=Allrow1_new and counter <=Allrow1_new:
    counter+=1
    void_runner = ws3.cell(index_rm_void,6).value
    if void_runner == None:
#        ws3.delete_rows(index_rm_void,1)
        print(index_rm_void)
#        index_rm_void-=1
    index_rm_void+=1
wb3.save(dir_save_C+"resultC/"+file_name_C)
'''
