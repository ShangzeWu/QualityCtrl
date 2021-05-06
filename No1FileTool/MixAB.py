# -*- coding: UTF-8 -*-
from openpyxl import *
import os
import shutil

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print(113344422+len(file_lists))
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/QualityCtrl/No1FileTool"
#print(path)
dir_A = path+'/resultA/' #用来读取A文件 的 路径
dir_B = path+'/resultB/' #用来读取B文件 的 路径
save_dir = path+'/MixAB/' #用来保存文件 的 路径

file_name_A = find_new_file(dir_A)
file_name_B = find_new_file(dir_B)

shutil.rmtree('MixAB')
os.mkdir('MixAB')
shutil.rmtree('MixAB_A')
os.mkdir('MixAB_A')
shutil.rmtree('MixAB_B')
os.mkdir('MixAB_B')

#业务逻辑
wb1 = load_workbook(dir_A+file_name_A)
ws1 = wb1[wb1.sheetnames[0]]
wb2 = load_workbook(dir_B+file_name_B)
ws2 = wb2[wb2.sheetnames[0]]

Allrow1 = ws1.max_row
Allcol1 = ws1.max_column
Allrow2 = ws2.max_row
Allcol2 = ws2.max_column
counter = 0
index_A_col=1
while index_A_col<=Allcol1:
        if ws1.cell(1,index_A_col).value=="运单编号":
#               print("check1")
                index_B_col=1
                while index_B_col<=Allcol2:
                        if ws2.cell(1,index_B_col).value=="运单号":
#                               print("check2")
                                index_A_row=2
                                while index_A_row<=Allrow1:
                                        value_A = ws1.cell(index_A_row,index_A_col).value
                                        if value_A == None:
                                            index_A_row=index_A_row+1
                                            continue
                                        else:
                                            value_A = int(value_A)
                                            index_B_row=2
                                            while index_B_row<=Allrow2:
                                                value_B = ws2.cell(index_B_row,index_B_col).value
                                                if value_B == None:
                                                    index_B_row = index_B_row+1
                                                    continue
                                                else:
                                                    value_B = int(value_B)
                                                    if value_A == value_B:
                                                        if ws1.cell(index_A_row,index_A_col+5).value==None:
                                                            ws1.delete_rows(index_A_row,1) #A表删除一行
                                                            index_A_row = index_A_row-1
                                                            print(value_A)
                                                        else:
                                                            ws2.delete_rows(index_B_row,1) #B表删除一行
                                                    index_B_row = index_B_row+1
                                            index_A_row = index_A_row+1
                        index_B_col=index_B_col+1
        index_A_col=index_A_col+1
wb1.save(path+"/MixAB_A/"+file_name_A)
wb2.save(path+"/MixAB_B/"+file_name_B)

wb1 = load_workbook(path+'/MixAB_A/'+file_name_A)
wb2 = load_workbook(path+'/MixAB_B/'+file_name_B)
wb1.save(save_dir+file_name_A)
wb2.save(save_dir+file_name_B)


