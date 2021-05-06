# -*- coding: UTF-8 -*-
from openpyxl import *
#from openpyxl.styles.numbers import FORMAT_TEXT
from datetime import datetime
import os
import time

time_start = time.time()

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/QualityCtrl/No1FileTool"
#print(path)
dir = path+'/uploadB/' #用来读取文件 的 路径
save_dir = path+'/resultB/' #用来保存文件 的 路径

Filename = find_new_file(dir)
wb = load_workbook(dir+Filename)
ws = wb[wb.sheetnames[0]]

Allrow = ws.max_row
Allcol = ws.max_column

index_col=1
while index_col<=Allcol:
#	print(ws.cell(1,index_col).value)
	if ws.cell(1,index_col).value=="包号"or ws.cell(1,index_col).value=="扫描网点"or ws.cell(1,index_col).value=="扫描类型"or ws.cell(1,index_col).value=="上传时间"or ws.cell(1,index_col).value=="扫描员"or ws.cell(1,index_col).value=="扫描员编号"or ws.cell(1,index_col).value=="收/派件员"or ws.cell(1,index_col).value=="上/下一站"or ws.cell(1,index_col).value=="重量"or ws.cell(1,index_col).value=="物品类型"or ws.cell(1,index_col).value=="寄件网点"or ws.cell(1,index_col).value=="寄件客户"or ws.cell(1,index_col).value=="问题件标识"or ws.cell(1,index_col).value=="任务号/车牌号"or ws.cell(1,index_col).value=="停滞时长":
		ws.delete_cols(index_col,1)
		index_col=index_col-1
		Allcol=Allcol-1
	index_col=index_col+1

#print(type(ws.cell(2,2)))
#print(type(ws.cell(2,2).value))
#print(ws.cell(2,2).value)

wb.save(save_dir+Filename)
#标注一次到件和二次到件

format_pattern = '%Y-%m-%d %H:%M:%S'
cur_time = datetime.now()
# 将 'cur_time' 类型时间通过格式化模式转换为 'str' 时间
cur_time = cur_time.strftime(format_pattern)
#print(cur_time)
cur_date = cur_time[0:10]+" 06:00:00"

wb1 = load_workbook(save_dir+Filename)
ws1 = wb1[wb1.sheetnames[0]]

allrows = ws1.max_row
allclos = ws1.max_column

index_row = 2
while index_row<=allrows:
	date_time = ws1.cell(index_row,2).value
	if date_time == None:
		index_row+=1
		continue
	else:
		date_time = str(date_time)
		
#		date1 = date_time[0:10]
#		time1 = date_time[11:]
#		print(date1)
#		print(time1)
		index_row+=1

wb1.save(save_dir+Filename)

#print('文件保存在'+save_dir+find_new_file(dir))
time_end=time.time()
print(time_end-time_start)
