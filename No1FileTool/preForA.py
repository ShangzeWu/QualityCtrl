# -*- coding: UTF-8 -*-
from openpyxl import *
import os
import time

time_start = time.time()

def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
    #print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
    #print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/QualityCtrl/No1FileTool"
#print(path)
dir = path+'/uploadA/' #用来读取文件 的 路径
save_dir = path+'/resultA/' #用来保存文件 的 路径
file_name = find_new_file(dir)

wb = load_workbook(dir+file_name)
ws = wb[wb.sheetnames[0]]
str1 = "江苏盐城"

Allrow = ws.max_row
Allcol = ws.max_column


#删除不需要的列
for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
	for cell in rows:
#		print('cell %s %s' % (cell.coordinate,cell.value))
		if cell.value=="电商平台"or cell.value=="中转站" or cell.value=="派送网点所属片区"or cell.value=='派送网点所属省份'or cell.value=='派送网点所属城市'or cell.value=='派送网点'or cell.value=='派送网点类型'or cell.value=='派送网点所属网点'or cell.value=='派件员'or cell.value=='中转发件时间'or cell.value=='网点发件时间'or cell.value=='派件时间'or cell.value=='派件入库时间'or cell.value=='签收时间'or cell.value=='签收入库时间'or cell.value=='签收时长/h'or cell.value=='签收时限'or cell.value=='当天签收时限'or cell.value=='签收渠道'or cell.value=='生鲜件标识'or cell.value=='共配件标识'or cell.value=='错分件原因'or cell.value=='错分件标识'or cell.value=='错分件登记网点'or cell.value=='错分件登记时间'or cell.value=='退回件扫描网点'or cell.value=='时效类型'or cell.value=='派件标识'or cell.value=='签收标识'or cell.value=='正常签收标识'or cell.value=='当天签收标识'or cell.value=='预售类型':
			ws.delete_cols(cell.column, 1)



#添加一列字段
ws.cell(row=1,column=7,value="下一站对应的三段码")
for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
	for cell in rows:
		#根据“下一站”字段，删除不需要的行
		if cell.value=="下一站":
			count_col = cell.column #记录当前的列数
			for cols in ws.iter_rows(min_row=2,min_col=count_col,max_col=count_col):
				for c_cell in cols:
					count_row = c_cell.row #记录当前的行数
					if c_cell.value=="江苏盐城集散中心":
						ws.delete_rows(count_row,1) #删除一行
						ws.insert_rows(count_row)
					if c_cell.value=="江苏盐城东沟镇分部":
						ws.cell(row=count_row,column=count_col+3,value=709)
					if c_cell.value=="江苏盐城益林镇分部":
						ws.cell(row=count_row,column=count_col+3,value=710)
					if c_cell.value=="江苏盐城大纵湖营业部":
						ws.cell(row=count_row,column=count_col+3,value=708)
					if c_cell.value=="江苏盐城龙冈营业部":
						ws.cell(row=count_row,column=count_col+3,value=523)
					if c_cell.value=="江苏盐城潘黄营业部":
						ws.cell(row=count_row,column=count_col+3,value=512)
					if c_cell.value=="江苏盐城黄尖营业部":
						ws.cell(row=count_row,column=count_col+3,value=707)
					if c_cell.value=="江苏盐城北蒋营业部" or c_cell.value=="江苏盐城秦南营业部":
						ws.cell(row=count_row,column=count_col+3,value=704)
					if c_cell.value=="江苏盐城盐东营业部":
						ws.cell(row=count_row,column=count_col+3,value=707)
		#根据“签收网点”字段，删除不需要的行
		if cell.value=="签收网点":
			count_col = cell.column#记录当前的列数
			for cols in ws.iter_rows(min_row=2,min_col=count_col,max_col=count_col):
				for c_cell in cols:
					str2 = c_cell.value
					if str2 == None:
						continue
					count_row = c_cell.row #记录当前的行数
					if str2=="申通集团":   #删除带有“申通集团” 的行
						ws.delete_rows(count_row,1)
						ws.insert_rows(count_row)
					if  str1 in str2:     # 删除带有“江苏盐城···”的行
						ws.delete_rows(count_row,1)
						ws.insert_rows(count_row)
#去除空行
i=1
while i <= Allrow:
	if ws.cell(i,1).value == None:
		ws.delete_rows(i,1)
		i=i-1
		Allrow=Allrow-1
	i=i+1

wb.save(save_dir+file_name) # 保存文件的路径+文件的名字
time_end=time.time()
print(time_end-time_start)
